
# lIBRERIAS
from flask import Flask, render_template, Response, request, session, logging, redirect, url_for, flash , make_response # clase para tener una conexión
from flask_bootstrap import Bootstrap
from openpyxl import load_workbook, workbook  # Libreria excel
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill # estilos libreria excel
import pandas as pd
import smtplib, getpass, os, re, pyodbc, pdfkit, datetime, itertools, xlwt, bcrypt
from itertools import *
from itertools import zip_longest
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.encoders import encode_base64
import locale
import re # validar correos

import hashlib

#--------------------------------------------------
app = Flask(__name__)
# configuración de la BD en Access

connStr = (
    r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
    r"DBQ=D:\\workspace_flask\\AMS\\amsdb.accdb;"
)

cnxn = pyodbc.connect(connStr)
# cur para conexión a BD y ejecutar consultas
cur = cnxn.cursor()
#settings

app.secret_key = 'mysecretkey'

# definimos con @app.route las rutas o link de acceso 

#CODIGO DE RUTAS PARA ADMIN 

@app.route('/', methods=['GET','POST']) 
def index ():
    if request.method == 'GET':
        #return render_template ("login.html")
        return render_template ("login.html") 
    else:
        #return render_template ("login.html")
        return render_template ("login.html") 
    
@app.route('/login', methods=['GET','POST']) # ruta que mostrara la pagina del login
def login():
    print("request.method ::: ",request.method)  
    if request.method == 'GET':
        return render_template ("login.html")
    elif request.method =='POST':
        email = request.form['email']
        password_login = request.form['password'] #.encode('utf-8')
       
       
        passw = hashlib.md5(password_login.encode())
        print("PASSWORD : ",  password_login)
        print("PASSWORD MD5 : ",  passw.hexdigest())

        cur = cnxn.cursor()
        cur.execute ("SELECT * FROM users WHERE email = ? and password = ?", (email, passw.hexdigest()))
        user = cur.fetchone()

        print("VA ACA.... ::: ", user)

        # cur.close()
        if user != None: 
            print("VA ACA.... 1 ::: ")
  
            session['nombre'] = (user[1]) #[0] es la posición en BD de la tabla user que quiero que registre en la tabla. 
            print("VA ACA.... ::: ")
               
            session['email'] = user[2]
            cur.execute(u'INSERT INTO tblogin (email) VALUES (?)', (email,))
            print(email)
           
            cur.commit()
            return render_template ("home.html")
            #return render_template ("admin.html")
        else:
            return render_template("login.html", error="Usuario y/o constraseña son erroneos")


@app.route('/reg_user', methods=['GET','POST']) # 
def reg_user ():
    if request.method == 'GET':
       return render_template ("reg_user.html")
    else: 
        iduser=request.form['iduser']
        nombre = request.form['nombre']
        email = request.form['email']
        password = request.form['password']
        hash_password = hashlib.md5(password.encode()) # para encriptar contraseña
        #cur = mysql.connection.cursor()
        cur = cnxn.cursor()
        cur.execute(u'INSERT INTO users (iduser, nombre, email, password) VALUES (?, ?, ?, ?)', 
        (iduser, nombre, email, hash_password.hexdigest()))
       
        cur.commit()
        session['nombre'] = nombre
        session['email'] = emailiduser = request.form['iduser']
        return render_template ("login.html")

@app.route('/home', methods=['GET','POST']) # 
def home():
    if request.method == 'GET':        
        return render_template ("home.html")
    else:
        return render_template ("home.html")
        
@app.route('/dashboard', methods=['GET','POST']) # 
def dashboard():
    if request.method == 'GET':        
        return render_template ("dashboard.html")




# CÓDIGO MODULO ADMINISTRATIVO 

@app.route('/admin', methods=['GET','POST'])  # 
def admin ():
    if request.method == 'GET':
        return render_template("admin.html")
    else:
        return render_template("admin.html")

@app.route('/etapas', methods=['GET','POST'])  # 
def etapas ():
    if request.method == 'GET':
       return render_template ("etapas.html")
    else:
       return render_template ("etapas.html")

@app.route('/crear_etapas', methods=['GET','POST']) # 
def crear_etapas():
    if request.method == 'GET':
        return render_template ("etapas.html") 
    else:
        nombre_etapa = request.form['nombre_etapa']
        email = session['email']
        if (len(nombre_etapa)>=1):
            #cur = mysql.connection.cursor()
            cur = cnxn.cursor()
            cur.execute(u'INSERT INTO etapas (nom_etapa, nom_usuario) VALUES (?, ?)', 
            (nombre_etapa, email,))
            #mysql.connection.commit()
            cur.commit()
            print (nombre_etapa)
            print (email)
            flash('Datos Guardados exitosamente')
            return render_template ("etapas.html")

@app.route('/mostrar_etapas', methods=['GET','POST']) # 
def mostrar_etapas():
    #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur = cnxn.cursor()
    cur.execute('SELECT id, nom_etapa, nom_usuario FROM etapas')
    result = cur.fetchall()
    total=len(result)
    return render_template("etapas.html", result = result, total=total)

@app.route('/update_etapa', methods = ['GET', 'POST'])
def update_etapa():
    if request.method == 'POST':
        id = request.form['id']
        print(id)
        nom_etapa = request.form['nom_etapa']     
        print(nom_etapa)
        cur = cnxn.cursor()
        cur.execute('UPDATE etapas SET nom_etapa=? WHERE id=?', nom_etapa, id)
        cur.commit()
        flash("Employee Updated Successfully")
        return redirect(url_for('etapas'))

@app.route('/delete_etapa/<id>/', methods = ['GET', 'POST'])
def delete_etapa(id):
        cur = cnxn.cursor()
        cur.execute('DELETE FROM etapas WHERE id=?', id)
        cur.commit()
        flash("Registro eliminado exitosamente")
        return redirect(url_for('etapas'))





#modulo de componentes
@app.route('/componentes', methods=['GET','POST'])
def componentes():
    if request.method == 'POST':       
        if request.form.get('Clic') == 'Clic':
            #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            cur = cnxn.cursor()
            cur.execute('SELECT nom_etapa, nom_usuario FROM etapas')
            result = cur.fetchall()
            print("Encrypted")
            return render_template("componentes.html", result=result)
        elif  request.form.get('Guardar') == 'Guardar':
            seleccione_etapa = request.form['seleccione_etapa']
            nom_componente = request.form['nom_componente']
            email = session['email']
            #cur = mysql.connection.cursor()
            cur = cnxn.cursor()
            cur.execute(u'INSERT INTO componentes (nombre, nom_usuario, nom_etapa) VALUES (?, ?, ?)', 
            (nom_componente , email,  seleccione_etapa))
            #mysql.connection.commit()
            cur.commit()
            print ( nom_componente )
            print (email)
            print (seleccione_etapa)
            flash('Datos Guardados exitosamente')
                            # pass # do something else
            print("Decrypted")
            return render_template("componentes.html")
        else:
                # pass # unknown
            return render_template("componentes.html")
    elif request.method == 'GET':
            # return render_template("index.html")
        print("No Post Back Call")
        return render_template("componentes.html")

@app.route('/mostrar_componentes', methods=['GET','POST'])
def mostrar_componentes():
    #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur = cnxn.cursor()
    cur.execute('SELECT id, nombre, nom_etapa, nom_usuario FROM componentes')
    result_comp = cur.fetchall()
    total = len(result_comp)
    print(total)
    return render_template("componentes.html", result_comp = result_comp, total=total) 

@app.route('/update_comp', methods = ['GET', 'POST'])
def update_comp():
    if request.method == 'POST':
        id = request.form['id']
        print(id)
        nombre = request.form['nombre']     
        cur = cnxn.cursor()
        cur.execute('UPDATE componentes SET nombre=? WHERE id=?', nombre, id)
        cur.commit()
        flash("Employee Updated Successfully")
        return redirect(url_for('componentes'))

@app.route('/delete_comp/<id>/', methods = ['GET', 'POST'])
def delete_comp(id):
    cur = cnxn.cursor()
    cur.execute('DELETE FROM componentes WHERE id=?', id)
    cur.commit()
    flash("Registro eliminado exitosamente")
    return redirect(url_for('componentes'))





# Modulo de Estados
@app.route('/estados', methods=['GET','POST']) # 
def estados():
    if request.method == 'GET':
       return render_template ("estados.html")
    else:
        return render_template ("estados.html")
    
@app.route('/crear_estados', methods=['GET','POST']) # 
def crear_estados():
    if request.method == 'GET':
        return render_template ("estados.html") 
    else:
        nom_estado = request.form['nom_estado']
        email = session['email']
        if (len(nom_estado)>=1):
            #cur = mysql.connection.cursor()    
            cur = cnxn.cursor()
            cur.execute(u'INSERT INTO estados (nombre, nom_usuario) VALUES (?, ?)', 
            (nom_estado, email))
            #mysql.connection.commit()
            cur.commit()
            print (nom_estado)
            print (email)
            flash('Datos Guardados exitosamente')
            return render_template ("estados.html")

@app.route('/mostrar_estados', methods=['GET','POST']) # 
def editar_estados():
    #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur = cnxn.cursor()
    cur.execute('SELECT id, nombre, nom_usuario FROM estados')
    result = cur.fetchall()
    total = len(result)
    return render_template("estados.html", result = result, total=total)

@app.route('/update_estado', methods = ['GET', 'POST'])
def update_estado():
    if request.method == 'POST':
        id = request.form['id']
        print(id)
        nombre = request.form['nombre']     
        cur = cnxn.cursor()
        cur.execute('UPDATE estados SET nombre=? WHERE id=?', nombre, id)
        cur.commit()
        flash("Employee Updated Successfully")
        return redirect(url_for('estados'))

@app.route('/delete_estado/<id>/', methods = ['GET', 'POST'])
def delete_estados(id):
    cur = cnxn.cursor()
    cur.execute('DELETE FROM estados WHERE id=?', id)
    cur.commit()
    flash("Registro eliminado exitosamente")
    return redirect(url_for('estados'))




# Modulo de Sistemas
@app.route('/sistemas', methods=['GET','POST']) #
def sistemas():
    if request.method == 'GET':
       return render_template ("sistemas.html")                                  
    else:
       return render_template ("sistemas.html")                                  
        
@app.route('/crear_sistemas', methods=['GET','POST']) # 
def crear_sistemas():
    if request.method == 'GET':
        return render_template ("sistemas.html") 
    else:
        email = session['email']
        tipo_sistema = request.form['tipo_sistema']
        num_sistema = request.form['num_sistema']
        nom_sistema = tipo_sistema + ' # ' + num_sistema
        if (len(tipo_sistema)>=1) and (len(num_sistema)>=1):
            #cur = mysql.connection.cursor()    
            cur = cnxn.cursor()
            cur.execute(u'INSERT INTO sistemas (tipo, numero, nombre, nom_usuario ) VALUES (?, ?, ?, ?)', 
            (tipo_sistema, num_sistema, nom_sistema, email))
            #mysql.connection.commit()
            cur.commit()
            flash('Datos Guardados exitosamente')
            return render_template ("sistemas.html")

@app.route('/mostrar_sistemas', methods=['GET','POST']) #
def mostrar_sistemas():
    #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur = cnxn.cursor()
    cur.execute('SELECT id, tipo, numero, nombre, nom_usuario FROM sistemas')
    result = cur.fetchall()
    total = len(result)
    return render_template("sistemas.html", result = result, total=total)

@app.route('/update_sistema', methods = ['GET', 'POST'])
def update_sistema():
    if request.method == 'POST':
        id = request.form['id']
        print(id)
        tipo = request.form['tipo']  
        numero = request.form['numero']
        nombre = tipo + ' # ' + numero 
        cur = cnxn.cursor()
        cur.execute('UPDATE sistemas SET tipo=?, numero=?, nombre=? WHERE id=?', tipo, numero, nombre, id)
        cur.commit()
        flash("Employee Updated Successfully")
        return redirect(url_for('sistemas'))

@app.route('/delete_sistema/<id>/', methods = ['GET', 'POST'])
def delete_sistema(id):
    cur = cnxn.cursor()
    cur.execute('DELETE FROM sistemas WHERE id=?', id)
    cur.commit()
    flash("Registro eliminado exitosamente")
    return redirect(url_for('sistemas'))





# Modulo de Pozos
@app.route('/pozos', methods=['GET','POST'])
def pozos():
    if request.method == 'POST':
        if request.form.get('Clic') =='Ver|Sistema':
           #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
           cur = cnxn.cursor()
           cur.execute('SELECT nombre, nom_usuario FROM sistemas')
           result = cur.fetchall()
           print("Encrypted")
           return render_template("pozos.html", result=result)
        elif  request.form.get('Guardar') == 'Guardar':
            seleccione_sistema = request.form['seleccione_sistema']
            nom_pozo = request.form['nom_pozo']
            ubicacion_pozo = request.form['ubicacion_pozo']
            email = session['email']
            #cur = mysql.connection.cursor()
            cur = cnxn.cursor()
            cur.execute(u'INSERT INTO tb_pozos (nombre, ubicacion, sistemas_nombre, nombre_usuario) VALUES (?, ?, ?, ?)', 
            (nom_pozo, ubicacion_pozo, seleccione_sistema, email))
           # mysql.connection.commit()
            cur.commit() 
            flash('Datos Guardados exitosamente')
                            # pass # do something else
            print("Decrypted")
            return render_template("pozos.html")
        else:
                # pass # unknown
            return render_template("pozos.html")
    elif request.method == 'GET': 
            # return render_template("index.html")
        print("No Post Back Call")
        return render_template("pozos.html")
    
@app.route('/mostrar_pozos', methods=['GET','POST'])
def mostrar_pozos():
    #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur = cnxn.cursor()
    cur.execute('SELECT id, nombre, ubicacion, sistemas_nombre, nombre_usuario FROM tb_pozos')
    result_pozo = cur.fetchall()
    total = len(result_pozo)
    return render_template("pozos.html", result_pozo = result_pozo, total=total) 
        
@app.route('/update_pozo', methods = ['GET', 'POST'])
def update_pozo():
    if request.method == 'POST':
        id = request.form['id']
        print(id)
        nombre = request.form['nombre']     
        cur = cnxn.cursor()
        cur.execute('UPDATE tb_pozos SET nombre=? WHERE id=?', nombre, id)
        cur.commit()
        flash("Employee Updated Successfully")
        return redirect(url_for('pozos'))

@app.route('/delete_pozo/<id>/', methods = ['GET', 'POST'])
def delete_pozo(id):
    cur = cnxn.cursor()
    cur.execute('DELETE FROM tb_pozos WHERE id=?', id)
    cur.commit()
    flash("Registro eliminado exitosamente")
    return redirect(url_for('pozos'))





# Modulo de creación de proyectos        
@app.route('/proyectos', methods=['GET','POST'])
def proyectos():
    if request.method == 'POST':
        if request.form.get('Clic') == 'Ver opciones':
            #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            cur = cnxn.cursor()
            cur.execute('SELECT nombre, sistemas_nombre FROM tb_pozos')
            result = cur.fetchall()
            print("Encrypted")
            return render_template ("proyectos.html", result = result)
        elif  request.form.get('Guardar') == 'Guardar': 
            sistemas = request.form['sist']
            seleccione_pozo = request.form['seleccione_pozo']
            nom_proyecto = seleccione_pozo+"_"+sistemas
            fecha_inicial = request.form['fecha_inicial']
            if len(fecha_inicial)==0:
                fecha_inicial ='AAAA-MM-DD'
            fecha_estimada_fin = request.form['fecha_estimada_fin']
            if len(fecha_estimada_fin)==0:
                fecha_estimada_fin ='AAAA-MM-DD'
            fecha_real_fin = request.form['fecha_real_fin']
            if len(fecha_real_fin)==0:
                fecha_real_fin ='AAAA-MM-DD'           
            Descripcion = request.form['Descripcion']
            email = session['email']
                #cur = mysql.connection.cursor()
            cur = cnxn.cursor()
            cur.execute(u'INSERT INTO proyectos (nombre,  fecha_inicio, fecha_estimada_fin, fecha_real_fin, nom_pozo, nom_sistema, descripcion, nom_usuario) VALUES (?, ?, ?, ?, ?, ?, ?, ?)', 
            (nom_proyecto,  fecha_inicial, fecha_estimada_fin, fecha_real_fin, seleccione_pozo, sistemas, Descripcion, email))
                            #mysql.connection.commit()
            cur.commit()
            flash('Datos Guardados exitosamente')
            return render_template ("proyectos.html") 
        else:
                    # pass # unknown
            return render_template("proyectos.html")
    elif request.method == 'GET':
            # return render_template("index.html")
        print("No Post Back Call")
        return render_template("proyectos.html")

@app.route('/mostrar_proyectos', methods=['GET','POST'])
def mostrar_proyectos():
    #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur = cnxn.cursor()
    cur.execute('SELECT idProyectos,  nombre, descripcion, nom_pozo, nom_sistema, fecha_inicio, fecha_estimada_fin, fecha_real_fin, descripcion,  nom_usuario FROM proyectos')
    result_proy = cur.fetchall()
    total = len(result_proy)
    return render_template("proyectos.html", result_proy = result_proy, total=total) 

@app.route('/update_proyecto', methods = ['GET', 'POST'])
def update_proyecto():
    if request.method == 'POST':
        id = request.form['id']
        fecha_i = request.form['fecha_inicial']   
        fecha_e = request.form['fecha_estimada_fin']   
        fecha_f = request.form['fecha_real_fin'] 
        descripcion = request.form['Descripcion']  
          
        cur = cnxn.cursor()
        cur.execute('UPDATE proyectos SET fecha_inicio=?, fecha_estimada_fin=?, fecha_real_fin=?, descripcion=? WHERE idProyectos=?',  fecha_i, fecha_e, fecha_f, descripcion, id)
        cur.commit()
        flash("Employee Updated Successfully")
        return redirect(url_for('proyectos'))

@app.route('/delete_proyecto/<id>/', methods = ['GET', 'POST'])
def delete_proyecto(id):
    cur = cnxn.cursor()
    cur.execute('DELETE FROM proyectos WHERE idproyectos=?', id)
    cur.commit()
    flash("Registro eliminado exitosamente")
    return redirect(url_for('proyectos'))





# Modulo de creación de usuario
@app.route('/user_admin', methods=['GET','POST']) # 
def user_admin():
    if request.method == 'GET':
       return render_template ("user_admin.html")                                  
    else:
       return render_template ("user_admin.html")                                  
        
@app.route('/crear_user_admin', methods=['GET','POST']) # 
def crear_user_admin():
    if request.method == 'POST':
        if  request.form.get('Guardar') == 'Crear':
            iduser= request.form['iduser']
            nombre = request.form['nombre']
            rol = "Inactivo"
            email = request.form['email']
            if re.match('^[(a-z0-9\_\-\.)]+@[(a-z0-9\_\-\.)]+\.[(a-z)]{2,15}$',email.lower()):
                print ("Correo correcto")
                password = request.form['password'].encode('utf-8')
                hash_password = bcrypt.hashpw(password, bcrypt.gensalt()) # para encriptar contraseña
                # cur = mysql.connection.cursor()
                cur = cnxn.cursor()       
                cur.execute(u'INSERT INTO users (iduser, nombre, rol, email, password) VALUES (?, ?, ?, ?, ?)', 
                (iduser, nombre, rol, email, hash_password))
                #mysql.connection.commit()
                cur.commit()
                session['nombre'] = nombre
                session['email'] = emailiduser = request.form['iduser']
                flash('Datos Guardados exitosamente')
            return render_template ("user_admin.html")
        else:# pass # unknown
            return render_template("user_admin.html")
    
    elif request.method == 'GET': 
            # return render_template("index.html")
        print("No Post Back Call")
        return render_template("user_admin.html")

@app.route('/mostrar_user_admin', methods=['GET','POST']) # 
def mostrar_user_admin():
    #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur = cnxn.cursor()
    cur.execute('SELECT iduser, nombre, email, rol FROM users')
    result_user = cur.fetchall()
    total = len(result_user)
    return render_template("user_admin.html", result_user = result_user, total=total)

@app.route('/update_usuario', methods = ['GET', 'POST'])
def update_usuario():
    if request.method == 'POST':
        id = request.form['id']
        nombre = request.form['nombre']   
        rol = request.form['rol']   
        email = request.form['email'] 
        cur = cnxn.cursor()
        cur.execute('UPDATE users SET nombre=?, rol=?, email=? WHERE iduser=?', nombre, rol, email, id)
        cur.commit()
        flash("Employee Updated Successfully")
        return redirect(url_for('user_admin'))

@app.route('/delete_usuario/<id>/', methods = ['GET', 'POST'])
def delete_usuario(id):
    cur = cnxn.cursor()
    cur.execute('DELETE FROM users WHERE iduser=?', id)
    cur.commit()
    flash("Registro eliminado exitosamente")
    return redirect(url_for('user_admin'))


# Modulo de creación de rol
""" @app.route('/rol_user', methods=['GET','POST'])  # 
def rol_user ():
    if request.method == 'GET':
       return render_template ("rol_user.html")
    else:
       return render_template ("rol_user.html")

@app.route('/crear_rol', methods=['GET','POST']) # 
def crear_rol():
    if request.method == 'GET':
        return render_template ("rol_user.html") 
    else:
        nombre_rol = request.form['nombre_etapa']
        email = session['email']
        if (len(nombre_rol)>=1):
            #cur = mysql.connection.cursor()
            cur = cnxn.cursor()
            cur.execute(u'INSERT INTO rol (nombre, nom_usuario) VALUES (?, ?)', 
            (nombre, email))
            #mysql.connection.commit()
            cur.commit()
            flash('Datos Guardados exitosamente')
            return render_template ("rol_user.html")

@app.route('/editar_rol', methods=['GET','POST']) # 
def editar_rol():
    #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur = cnxn.cursor()
    cur.execute('SELECT nombre, nom_usuario FROM rol')
    result = cur.fetchall()
    return render_template("rol_user.html", result = result)

 """


# CÓDIGO MODULO PROYECTO 
# Modulo de Gestión de proyectos        
@app.route('/Proyecto', methods=['GET','POST'])
def Proyecto():
    if request.method == 'GET':
        return render_template ("Proyecto.html")

@app.route('/menu_proyecto', methods=['GET','POST'])
def menu_proyecto():
    if request.method == 'GET':
        return render_template ("menu_proyecto.html")
    else:
        return render_template ("menu_proyecto.html")
        
@app.route('/datos_proyecto', methods=['GET','POST'])
def datos_proyecto():
    if request.method == 'POST':
        if request.form.get('List') == 'Activar Opciones':
            #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            cur = cnxn.cursor()
            cur.execute('SELECT nombre FROM proyectos') 
            result_proyecto = cur.fetchall()
            cur.execute('SELECT DISTINCT nom_etapa FROM componentes') # etapas Distin sirve para que no aparezca duplicado el dato
            result_etapa = cur.fetchall()
            cur.execute("SELECT nombre FROM componentes") # etapas y componentes
            result_compo = cur.fetchall()
            cur.execute('SELECT nombre FROM estados')
            result_estado = cur.fetchall()
            cur.execute('SELECT abreviatura FROM monedas')
            result_moneda = cur.fetchall()
            return render_template("datos_proyecto.html", result_moneda=result_moneda, result_estado=result_estado, result_etapa=result_etapa , result_proyecto=result_proyecto, result_compo=result_compo)
        elif  request.form.get('Guardar') == 'Guardar Datos': 
            seleccione_proyecto = request.form['seleccione_proyecto']
            project = seleccione_proyecto
            separacion = project.split('_')
            pozo= separacion[0]
            sistema= separacion[1]
            print(separacion)
            print(pozo)
            print(sistema)
            seleccione_estado = request.form['seleccione_estado']
            fecha_inicial = request.form['fecha_inicial'] # date_available
            print(len(fecha_inicial))
            comentarios = request.form['comentarios']
            email = session['email']
            seleccione_etapa = request.form['seleccione_etapa']
            seleccione_componente = request.form['seleccione_componente']
            moneda = request.form['seleccione_abreviatura']
            Valor = request.form['valor']
            if len(Valor) >3:
                Valorf = format(int(Valor), ",") # separa por comas
                print(Valorf)
                valor_moneda = '$'+ '   '+Valorf +'      '+ moneda
            elif len (Valor)>0<3:
                valor_moneda = '$'+ '   '+Valor+'      '+ moneda   
            elif len(Valor) == 0:
                valor_moneda = "Sin Valor"
            
            if len(fecha_inicial)>0:
                fecha = fecha_inicial
            elif len(fecha_inicial) == 0:
                fecha = "AAAA-MM-DD"

            cur = cnxn.cursor()
            cur.execute(u'INSERT INTO resumen (nom_proyecto, nom_pozo, nom_sistema, nom_etapa, nom_componente, moneda, Valor, valor_moneda, nom_estado, date_available, comentarios,  nom_usuario) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)', 
            (seleccione_proyecto, pozo, sistema, seleccione_etapa, seleccione_componente, moneda, Valor, valor_moneda, seleccione_estado, fecha, comentarios, email))
            cur.commit()           
            flash('Datos Guardados exitosamente')
            return render_template("datos_proyecto.html")
        else:
             return render_template("datos_proyecto.html")
    elif request.method == 'GET':
            # return render_template("index.html")
        print("No Post Back Call")
        return render_template("datos_proyecto.html")

@app.route('/resumen_proyecto', methods=['GET','POST'])
def resumen_proyecto():

     #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur = cnxn.cursor()
    cur.execute('SELECT nombre FROM proyectos') 
    list_proyecto = cur.fetchall()
    cur.execute('SELECT nom_etapa FROM etapas') 
    list_etapa = cur.fetchall()

    if request.method == 'GET':
        # return render_template("index.html")
        print("No Post Back Call")
        return render_template("resumen_proyecto.html", list_proyecto=list_proyecto, list_etapa=list_etapa)
    
    if request.method == 'POST':
        
        if request.form.get('List') == 'Activar Opciones':
            #cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            cur = cnxn.cursor()
            cur.execute('SELECT nombre FROM proyectos') 
            list_proyecto = cur.fetchall()
            cur.execute('SELECT nom_etapa FROM etapas') 
            list_etapa = cur.fetchall()
            return render_template("resumen_proyecto.html", list_proyecto=list_proyecto, list_etapa=list_etapa ) 
        
        elif request.form.get('Consultar') == 'Consultar':
            selec_proyecto = request.form['selec_proyecto']
            seleccione_etapa = request.form['seleccione_etapa']
            pozo_sist = selec_proyecto.split('_')
            #print(pozo_sist)
            Pozo = pozo_sist[0]
            #print(Pozo)
            Sist = pozo_sist[1]
            #print(Sist)
            if seleccione_etapa  != 'Todas las Etapas':
                cur = cnxn.cursor()
                cur.execute("SELECT DISTINCT  nom_etapa FROM resumen WHERE nom_proyecto= ? and nom_etapa=?", (selec_proyecto, seleccione_etapa,))
                result_etapa =cur.fetchall()
                cur.execute("SELECT id,  nom_etapa, nom_componente, Valor, valor_moneda, nom_estado, date_available, comentarios, fecha_registro, nom_usuario FROM resumen WHERE nom_proyecto= ? and nom_etapa=?", (selec_proyecto, seleccione_etapa,))
                result_proyecto2 = cur.fetchall()
                return render_template("resumen_proyecto.html", Pozo=Pozo, Sist=Sist,  selec_proyecto=selec_proyecto, seleccione_etapa=seleccione_etapa, result_etapa=result_etapa, result_proyecto2=result_proyecto2 )      
            if seleccione_etapa =='Todas las Etapas':
                cur = cnxn.cursor()
                cur.execute("SELECT DISTINCT  nom_etapa FROM resumen WHERE nom_proyecto= ?", (selec_proyecto,))
                result_etapa =cur.fetchall()
                cur.execute("SELECT id, nom_proyecto, nom_etapa, nom_componente, Valor, valor_moneda, nom_estado, date_available, comentarios, fecha_registro, nom_usuario FROM resumen WHERE nom_proyecto= ?", (selec_proyecto, ))
                result_proyecto2 = cur.fetchall()
                return render_template("resumen_proyecto.html", Pozo=Pozo, Sist=Sist ,  selec_proyecto=selec_proyecto, seleccione_etapa=seleccione_etapa, result_etapa=result_etapa, result_proyecto2=result_proyecto2 )       


@app.route('/update_resproyecto', methods = ['GET', 'POST'])
def update_resproyecto():
    if request.method == 'POST':
        id = request.form['id']
        valor = request.form['valor']
        estado = request.form['estado'] 
        date=request.form['date']
        comentarios=request.form['comentarios']
        if len(valor) >3:
            Valorf = format(int(valor), ",") # separa por comas
            print(Valorf)
            valor_moneda = '$'+ '   '+Valorf +'      '+ 'USD'
        elif len (valor)>0<3:
            valor_moneda = '$'+ '   '+valor+'      '+ 'USD'   
        elif len(valor) == 0:
            valor_moneda = "Sin Valor"
            
        if len(date)>0:
            fecha = date
        elif len(date) == 0:
            fecha = "AAAA-MM-DD"

        cur = cnxn.cursor()
        cur.execute('UPDATE resumen SET Valor=?, valor_moneda=?, nom_estado=?, date_available=?, comentarios=?  WHERE id=?', valor, valor_moneda, estado, fecha, comentarios, id)
        cur.commit()
        flash("Employee Updated Successfully")
        return render_template("resumen_proyecto.html")
        
#        return redirect(url_for('resumen_proyecto'))

@app.route('/delete_resproyecto/<id>/', methods = ['GET', 'POST'])
def delete_reproyecto(id):
    cur = cnxn.cursor()
    cur.execute('DELETE FROM resumen WHERE id=?', id)
    cur.commit()
    flash("Registro eliminado exitosamente")
    return redirect(url_for('resumen_proyecto'))

@app.route('/descarga_excel', methods=['GET','POST'])
def descarga_excel():
    if request.method == 'POST':         
        print ('ESTA BIEN')
        if request.form.get('Excel') == 'Excel':
          # codigo para generar el excel con los datos del resumen
          #  filesheet = "./Resumen.xlsx"  # variable para escribir la ruta en donde deseo guardar el archivo( carpeta templates) , y para crear el nombre del excel 
          # Codigo para cambiar el color a la celda y el tipo de letra 

            fuente_titulo_blanco = Font(size = 14, bold = True, color = "FFFFFF") 
            fuente_negrita_blanco = Font(size = 11, bold = True, color = "FFFFFF") 
            fuente_negrita_petroleo = Font(size = 11, bold = True, color = "000000") 
            fuente_normal_cursiva_petroleo = Font(size = 10, italic=True, color = "000000") 
            fondo_petroleo = PatternFill(fgColor="01454F", bgColor="01454F", fill_type="solid")
            fondo_gris = PatternFill(fgColor="CDCDCD", bgColor="CDCDCD", fill_type="solid")
            #fondo_gris_o = PatternFill(fgColor="9C9C9C", bgColor="9C9C9C", fill_type="solid")
            fondo_blanco = PatternFill(fgColor="FFFFFF", bgColor="FFFFFF", fill_type="solid")
            Pozo = request.form['pozo']
            print(Pozo)
            Sistema=request.form['sistema']
            print(Sistema)
            Etapa = request.form['etapa']
            print(Etapa)
            Proyecto= Pozo + '_' + Sistema
            print(Proyecto)           
            filesheet = "D:/workspace_flask/AMS/Resumen.xlsx"
            wb = Workbook()
            #wb = load_workbook(filesheet)
            sheet = wb.active 
            
            # Ajusto Ancho de columna ene excel
            
            sheet.merge_cells('A1:E1') # COMBINAR CELDAS
                
            #sheet.unmerge_cells('A1:G1') # Descombinar celdas
            
            #Coloreo las Celdas
      
            sheet['A1'].fill = fondo_petroleo
            sheet['B1'].fill = fondo_petroleo
            sheet['C1'].fill = fondo_petroleo   
            sheet['D1'].fill = fondo_petroleo
            sheet['E1'].fill = fondo_blanco
            sheet['F1'].fill = fondo_blanco
            sheet['G1'].fill = fondo_blanco
            
            sheet['A6'].fill = fondo_petroleo
            sheet['B6'].fill = fondo_petroleo
            sheet['C6'].fill = fondo_petroleo   
            sheet['D6'].fill = fondo_petroleo
            sheet['E6'].fill = fondo_petroleo
            sheet['F6'].fill = fondo_petroleo
            sheet['G6'].fill = fondo_petroleo
            sheet['H6'].fill = fondo_petroleo
            
            sheet['A2'].fill = fondo_blanco
            sheet['B2'].fill = fondo_blanco
            sheet['C2'].fill = fondo_blanco
            sheet['D2'].fill = fondo_blanco
            sheet['E2'].fill = fondo_blanco
            sheet['F2'].fill = fondo_blanco
            sheet['G2'].fill = fondo_blanco
            sheet['A3'].fill = fondo_blanco
            sheet['B3'].fill = fondo_blanco
            sheet['C3'].fill = fondo_blanco
            sheet['D3'].fill = fondo_blanco
            sheet['E3'].fill = fondo_blanco
            sheet['F3'].fill = fondo_blanco
            sheet['G3'].fill = fondo_blanco
            sheet['A4'].fill = fondo_blanco
            sheet['B4'].fill = fondo_blanco
            sheet['C4'].fill = fondo_blanco
            sheet['D4'].fill = fondo_blanco
            sheet['E4'].fill = fondo_blanco
            sheet['F4'].fill = fondo_blanco
            sheet['G4'].fill = fondo_blanco
            sheet['A5'].fill = fondo_blanco
            sheet['B5'].fill = fondo_blanco
            sheet['C5'].fill = fondo_blanco
            sheet['D5'].fill = fondo_blanco
            sheet['E5'].fill = fondo_blanco
            sheet['F5'].fill = fondo_blanco
            sheet['G5'].fill = fondo_blanco
            sheet['H1'].fill = fondo_blanco
            sheet['H2'].fill = fondo_blanco
            sheet['H3'].fill = fondo_blanco
            sheet['H4'].fill = fondo_blanco
            sheet['H5'].fill = fondo_blanco
                        
            
            # Aplico Fuentes a las Celdas
            sheet['A1'].font = fuente_titulo_blanco
            sheet['A2'].font = fuente_negrita_petroleo
            sheet['B2'].font = fuente_normal_cursiva_petroleo
            sheet['A3'].font = fuente_negrita_petroleo
            sheet['B3'].font = fuente_normal_cursiva_petroleo
            sheet['A4'].font = fuente_negrita_petroleo
            sheet['B4'].font = fuente_normal_cursiva_petroleo
            sheet['A5'].font = fuente_negrita_petroleo
            sheet['B5'].font = fuente_normal_cursiva_petroleo
            sheet['D3'].font = fuente_negrita_petroleo
            sheet['E3'].font = fuente_normal_cursiva_petroleo
            
            sheet['A6'].font = fuente_negrita_blanco
            sheet['B6'].font = fuente_negrita_blanco
            sheet['C6'].font = fuente_negrita_blanco  
            sheet['D6'].font = fuente_negrita_blanco
            sheet['E6'].font = fuente_negrita_blanco
            sheet['F6'].font = fuente_negrita_blanco
            sheet['G6'].font = fuente_negrita_blanco
            sheet['H6'].font = fuente_negrita_blanco

           # almaceno datos en celdas  
            sheet['A1'] = "INFORME RESUMEN DE PROYECTOS"
            sheet['A2'] = "Proyecto" 
            sheet['B2'] =  Proyecto
            sheet['A3'] = "Pozo"
            sheet['B3'] = Pozo
            sheet['A4'] = "Sistema"
            sheet['B4'] = Sistema
            sheet['A5'] = "Etapa"
            sheet['B5'] = Etapa
 
            # Consulta a Base de datos
            sheet['D3'] = "Fecha Reporte"
            sheet['E3'] = datetime.datetime.now()
            sheet['A6'] = "Etapa"
            sheet['B6'] = "Componente"
            sheet['C6'] = "Valor"
            sheet['D6'] = "Estado"
            sheet['E6'] = "Date Available"
            sheet['F6'] = "Comentarios"
            sheet['G6'] = "Fecha de Registro"
            sheet['H6'] = "Usuario que registro Datos" 
            
            if Etapa != 'Todas las Etapas':
                cur = cnxn.cursor()
                cur.execute('SELECT nom_etapa, nom_componente, Valor, nom_estado, date_available, comentarios, fecha_registro, nom_usuario FROM resumen WHERE nom_proyecto= ? and nom_etapa=?', (Proyecto, Etapa,)) 
                resultexcel = cur.fetchall()
                
                print(resultexcel)
                b=type(resultexcel)
                print (b)

                for rows in resultexcel: # for para insertar datos de la consulta en excel. 
                    listrows = list(rows)
                    sheet.append(listrows)
                        
                wb.save(filesheet)       
                flash('Reporte generado Exitosamente')

            if Etapa == 'Todas las Etapas':
               cur = cnxn.cursor()
               cur.execute('SELECT nom_etapa, nom_componente, Valor, nom_estado, date_available, comentarios, fecha_registro, nom_usuario FROM resumen WHERE nom_proyecto= ?', (Proyecto, )) 
               resultexcel = cur.fetchall()
              
               print(resultexcel)
               b=type(resultexcel)
               print (b)

               for rows in resultexcel: # for para insertar datos de la consulta en excel. 
                   listrows = list(rows)
                   sheet.append(listrows)
                      
               wb.save(filesheet)       
               flash('Reporte generado Exitosamente')

        return render_template("resumen_proyecto.html")
    else:
        print ('NO ESTA BIEN')
        return render_template("resumen_proyecto.html")         

 #Codigo para generar Informe, guardarlos en Pdf, y enviarlos por correo electronico

@app.route('/informe_proyecto', methods=['GET','POST'])
def informe_proyecto():
    if request.method == 'POST':
        if request.form.get('List') == 'Activar':
            cur = cnxn.cursor()
            cur.execute('SELECT nombre FROM proyectos') 
            list_proyecto = cur.fetchall()
            return render_template("informe_proyecto.html", list_proyecto=list_proyecto ) 
        elif  request.form.get('Generar') == 'Generar Informe':
            seleccione_proyecto = request.form['seleccione_proyecto']
            cur = cnxn.cursor()
            cur.execute("SELECT DISTINCT nom_pozo, nom_sistema FROM resumen WHERE nom_proyecto= ? ", (seleccione_proyecto,))
            result_pozo_sist =cur.fetchall()
            print(result_pozo_sist)
            cur.execute("SELECT DISTINCT nom_etapa FROM resumen WHERE nom_proyecto= ?", (seleccione_proyecto,))
            result_etapa =cur.fetchall()
            cur.execute("SELECT nom_etapa, nom_componente, Valor, nom_estado, date_available, comentarios, fecha_registro, nom_usuario FROM resumen WHERE nom_proyecto= ?", (seleccione_proyecto,))
            result_proyecto2 = cur.fetchall()
            ###Codigo generar PDF
            rendered =  render_template("informePDF.html",result_pozo_sist=result_pozo_sist, seleccione_proyecto=seleccione_proyecto, result_etapa=result_etapa, result_proyecto2=result_proyecto2 )
            
            path_wkhtmltopdf = r'D:\\workspace_flask\\AMS\\libs\\wkhtmltox\\bin\\wkhtmltopdf.exe' # ubicacion del programa para convertir html pdf
            
            config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf )
            pdf = str (pdfkit.from_string(rendered, "informe.pdf", configuration=config))
            response = make_response(pdf)
            flash('Informe Generado Exitosamente, Por favor revise su carpeta de descargas')
            return render_template("informe_proyecto.html", result_pozo_sist=result_pozo_sist, seleccione_proyecto=seleccione_proyecto, result_etapa=result_etapa, result_proyecto2=result_proyecto2 )      
        # codigo enviar informe por E-mail
        elif request.form.get('Enviar') =='Enviar Informe':
            print("**** Enviar email con Hotmail ****")
            user ="julicvt@outlook.es"
            password ="Soyexcelente$"

            #Para las cabeceras del email
            remitente = "julicvt@outlook.es"#input("From, ejemplo: administrador <admin@hotmail.com>: ")
            destinatario = ['julicvt@outlook.es','julicvt@gmail.com','leandres1985@gmail.com','ambiental.angy@gmail.com', 'jmanuelds89@gmail.com'] #input("To, ejemplo: amigo <amigo@mail.com>: ")
            asunto = "PRUEBA DE ENVIO DE INFORME DESDE APP AMS"
            mensaje = "Esto es una prueba, por favor omita este mensaje y eliminelo"
            archivo ="informe.pdf" 

            #Host y puerto SMTP de Hotmail
            hotmail = smtplib.SMTP('smtp.live.com', 25)

            #protocolo de cifrado de datos utilizado por Hotmail
            hotmail.starttls()

            #Credenciales
            hotmail.login(user, password)

            #muestra la depuración de la operacion de envío 1=true8                                                                                                                                                                                                                                                                                                                                                                          
            hotmail.set_debuglevel(1)

            header = MIMEMultipart()
            header['Subject'] = asunto
            header['From'] = remitente
            header['To'] = ','.join(destinatario)

            mensaje = MIMEText(mensaje, 'html') #Content-type:text/html
            header.attach(mensaje)

            if (os.path.isfile(archivo)):
                adjunto = MIMEBase('application', 'octet-stream')
                adjunto.set_payload(open(archivo, "rb").read())
                encode_base64(adjunto)
                adjunto.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(archivo))
                header.attach(adjunto)

            #Enviar email
            hotmail.sendmail(remitente, destinatario, header.as_string())

            #Cerrar la conexión SMTP
            hotmail.quit()       
            flash('Informe Enviado Exitosamente')
            return render_template("informe_proyecto.html")
        else:
            flash('Reporte NOOOOOOO generado Exitosamente')
            return render_template("informe_proyecto.html")     
    elif request.method == 'GET':
            # return render_template("index.html")
        print("No Post Back Call")
        return render_template("informe_proyecto.html")
    

#----------------------------------------------------------- 
if __name__ == '__main__':   #valida que si el archivo inicial corresponde a app.py, entonces corre el servidor en el puerto 3000
   app.secret_key = 'super secret key'
    #app.run(port = 3000, debug =True)
   app.run(debug=True)
